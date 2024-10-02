from telethon import TelegramClient
from telethon.errors import UsernameNotOccupiedError, UsernameInvalidError, FloodWaitError
from telethon.tl.types import User, Channel
from telethon.tl.types import UserStatusRecently, UserStatusLastMonth, UserStatusLastWeek, UserStatusOnline, UserStatusOffline
import openpyxl
import os
import logging
import asyncio
import random

# Список нескольких аккаунтов для распределения запросов
ACCOUNTS = [
    {'api_id': '20050740', 'api_hash': '8ec7885605b366fa23cbdf1d32071666', 'session': 'artem_osn'},
    {'api_id': '27207363', 'api_hash': 'b8f021aa452be5a06999377597e00823', 'session': 'mami31'},
    {'api_id': '25934933', 'api_hash': '5773f892bb97166d656dbe9a4b3889f8', 'session': 'tinkoff'},
    {'api_id': '24529865', 'api_hash': '8728b0f8db006c40d9fcb810ec447058', 'session': 'андрюха'},
    {'api_id': '23691283', 'api_hash': '07170919d82069a4b0d4b9b93e697301', 'session': 'ваня'},
    {'api_id': '28053099', 'api_hash': 'e68420e079ff462c09dda362176254ad', 'session': 'дед'},
    # Добавьте больше аккаунтов по мере необходимости
]

class Parser:
    def __init__(self, input_file, start_row=1):
        self.accounts = ACCOUNTS
        self.input_file = input_file
        self.output_file = 'output.xlsx'
        self.start_row = start_row  # С какого ряда начинать
        self.clients = []

    async def init_clients(self):
        # Инициализация клиентов Telegram для каждого аккаунта
        for account in self.accounts:
            client = TelegramClient(account['session'], account['api_id'], account['api_hash'])
            await client.start()
            self.clients.append((client, account['session']))  # Сохраняем кортеж (client, session)

    async def get_entity_safely(self, client, username):
        try:
            input_entity = await client.get_input_entity(username)
            entity = await client.get_entity(input_entity)
            return entity
        except (UsernameNotOccupiedError, UsernameInvalidError):
            return None
        except FloodWaitError as e:
            logging.warning(f"FloodWaitError: необходимо ожидание {e.seconds} секунд.")
            await asyncio.sleep(e.seconds * 1.5)  # Увеличенное время ожидания
            return await self.get_entity_safely(client, username)
        except Exception as e:
            logging.error(f"Ошибка при получении сущности для {username}: {e}")
            return None

    async def check_username(self, client, username):
        try:
            if not username:
                return "не существует", "N/A"

            entity = await self.get_entity_safely(client, username)

            if isinstance(entity, User):
                status = "человек"
                last_online = self.parse_user_status(entity.status)
            elif isinstance(entity, Channel):
                status = "канал" if entity.broadcast else "чат"
                last_online = "N/A"
            else:
                return "не существует", "N/A"

            return status, last_online

        except Exception as e:
            logging.error(f"Ошибка при проверке пользователя {username}: {e}")
            return "ошибка", "N/A"

    def parse_user_status(self, status):
        if isinstance(status, UserStatusRecently):
            return "недавно был в сети"
        elif isinstance(status, UserStatusLastWeek):
            return "был в сети на прошлой неделе"
        elif isinstance(status, UserStatusLastMonth):
            return "был в сети в прошлом месяце"
        elif isinstance(status, UserStatusOnline):
            return f"в сети ({status.expires.strftime('%Y-%м-%д %H:%M:%S')})"
        elif isinstance(status, UserStatusOffline):
            return f"был в сети {status.was_online.strftime('%Y-%м-%д %H:%М:%S')}"
        else:
            return "не видно"

    def get_last_processed_row(self):
        """Чтение выходного файла для получения последней обработанной строки."""
        if not os.path.exists(self.output_file):
            logging.info("Файл не существует, начинаем с первой строки.")
            return self.start_row  # Если выходного файла нет, начинаем с заданного start_row

        wb = openpyxl.load_workbook(self.output_file)
        sheet = wb.active

        # Проходим по строкам с конца и ищем последнюю непустую строку
        for row in range(sheet.max_row, 0, -1):
            username = sheet.cell(row=row, column=1).value  # Проверяем наличие значения в первой колонке (username)
            if username:  # Если имя пользователя существует в строке
                logging.info(f"Последний обработанный пользователь в строке: {row}. Начинаем с {row + 1}.")
                return row + 1  # Начинаем со следующей строки после последней записанной

        logging.info(f"Все строки пусты, начинаем с {self.start_row}.")
        return self.start_row  # Если все строки пусты, возвращаем стартовую строку

    async def process_usernames(self):
        try:
            await self.init_clients()  # Инициализация всех клиентов
            wb = openpyxl.load_workbook(self.input_file)
            sheet = wb.active

            # Создание или загрузка выходного Excel файла
            if os.path.exists(self.output_file):
                output_wb = openpyxl.load_workbook(self.output_file)
                output_sheet = output_wb.active
            else:
                output_wb = openpyxl.Workbook()
                output_sheet = output_wb.active
                output_sheet.append(['Username', 'Статус', 'Последний раз в сети', 'Используемый аккаунт'])

            # Определение последней обработанной строки
            last_processed_row = self.get_last_processed_row()
            logging.info(f"Начинаем проверку с строки: {last_processed_row}")

            # Собираем все имена пользователей, начиная с последней необработанной строки
            usernames = [row[0] for row in sheet.iter_rows(min_row=self.start_row, values_only=True) if row[0]]
            total_usernames = len(usernames)

            # Начинаем с последней обработанной строки
            for index, username in enumerate(usernames[last_processed_row:], start=last_processed_row):
                if username:
                    try:
                        client, session_name = random.choice(self.clients)  # Выбор случайного аккаунта для запроса
                        status, last_online = await self.check_username(client, username)
                        output_sheet.append([username, status, last_online, session_name])

                        # Вывод в консоль проверенного пользователя и текущей сессии
                        remaining = total_usernames - index
                        print(f"Проверен: {username}, Статус: {status}, Осталось: {remaining} пользователей, Аккаунт: {session_name}")

                        # Сохранение прогресса после каждых 500 пользователей
                        if index % 500 == 0:
                            output_wb.save(self.output_file)
                            print(f"Сохранение прогресса после {index} проверенных пользователей.")

                        # Пауза 5 минут после 500 пользователей
                        if index % 500 == 0:
                            print(f"Пауза на 5 минут после проверки {index} пользователей.")
                            await asyncio.sleep(1800)  # Пауза 5 минут

                        # Пауза 30 секунд после каждых 100 пользователей
                        if index % 100 == 0:
                            print(f"Пауза на 30 секунд после проверки {index} пользователей.")
                            await asyncio.sleep(600)  # Пауза 30 секунд

                        await asyncio.sleep(30)  # Увеличенная задержка между запросами (5 секунд)

                    except FloodWaitError as e:
                        if e.seconds > 300:
                            print(f"FloodWaitError: {session_name} заблокирован на {e.seconds} секунд. Смена аккаунта.")
                            self.clients.remove((client, session_name))  # Убираем заблокированный аккаунт
                            if not self.clients:
                                print(f"Все аккаунты заблокированы. Ожидание завершения блокировки {e.seconds}")
                                await asyncio.sleep(e.seconds)
                        else:
                            print(f"FloodWaitError: Ожидание {e.seconds} секунд.")
                            await asyncio.sleep(e.seconds)

                else:
                    logging.info("Пропуск пустой строки или пустого имени пользователя")

            # Сохранение выходного файла
            output_wb.save(self.output_file)
            return os.path.abspath(self.output_file)

        except Exception as e:
            logging.error(f"Ошибка при обработке списка пользователей: {e}")
            return None


    async def run(self):
        return await self.process_usernames()
